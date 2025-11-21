"""
Vistas personalizadas para el panel de docentes
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Avg, Count, Q
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Materia, Matricula, Calificacion, Asistencia, Notificacion
from accounts.models import CustomUser


def is_teacher(user):
    return user.role == 'docente'


@login_required
@user_passes_test(is_teacher)
def teacher_dashboard(request):
    """Dashboard principal para docentes"""
    # Materias del docente
    materias = Materia.objects.filter(docente=request.user, activa=True)

    # Estudiantes totales
    estudiantes_ids = Matricula.objects.filter(
        curso__materias__in=materias,
        activa=True
    ).values_list('estudiante_id', flat=True).distinct()
    total_estudiantes = len(estudiantes_ids)

    # Calificaciones recientes
    calificaciones_recientes = Calificacion.objects.filter(
        materia__in=materias
    ).select_related('estudiante', 'materia').order_by('-fecha_registro')[:10]

    # Estadísticas
    total_calificaciones = Calificacion.objects.filter(materia__in=materias).count()
    promedio_general = Calificacion.objects.filter(materia__in=materias).aggregate(Avg('nota'))['nota__avg'] or 0

    context = {
        'materias': materias,
        'total_estudiantes': total_estudiantes,
        'total_calificaciones': total_calificaciones,
        'promedio_general': round(promedio_general, 2),
        'calificaciones_recientes': calificaciones_recientes,
    }

    return render(request, 'teacher/dashboard.html', context)


# ==================== GESTIÓN DE CALIFICACIONES ====================

@login_required
@user_passes_test(is_teacher)
def calificaciones_lista(request):
    """Lista de calificaciones del docente"""
    materias = Materia.objects.filter(docente=request.user, activa=True)
    calificaciones = Calificacion.objects.filter(materia__in=materias).select_related(
        'estudiante', 'materia', 'materia__curso'
    ).order_by('-fecha_registro')

    # Filtros
    materia_id = request.GET.get('materia')
    periodo = request.GET.get('periodo')
    search = request.GET.get('search')

    if materia_id:
        calificaciones = calificaciones.filter(materia_id=materia_id)
    if periodo:
        calificaciones = calificaciones.filter(periodo=periodo)
    if search:
        calificaciones = calificaciones.filter(
            Q(estudiante__username__icontains=search) |
            Q(estudiante__first_name__icontains=search) |
            Q(estudiante__last_name__icontains=search)
        )

    context = {
        'calificaciones': calificaciones,
        'materias': materias,
        'periodos': Calificacion.PERIODO_CHOICES,
        'materia_filter': materia_id,
        'periodo_filter': periodo,
        'search_query': search,
    }
    return render(request, 'teacher/calificaciones_lista.html', context)


@login_required
@user_passes_test(is_teacher)
def calificacion_crear(request):
    """Crear nueva calificación"""
    materias = Materia.objects.filter(docente=request.user, activa=True)

    if request.method == 'POST':
        estudiante_id = request.POST.get('estudiante')
        materia_id = request.POST.get('materia')
        periodo = request.POST.get('periodo')
        nota = request.POST.get('nota')
        observaciones = request.POST.get('observaciones', '')

        # Verificar que la materia pertenece al docente
        materia = get_object_or_404(Materia, id=materia_id, docente=request.user)

        # Verificar si ya existe una calificación
        if Calificacion.objects.filter(estudiante_id=estudiante_id, materia=materia, periodo=periodo).exists():
            messages.error(request, 'Ya existe una calificación para este estudiante en este periodo')
            return redirect('teacher_calificacion_crear')

        # Crear calificación
        calificacion = Calificacion.objects.create(
            estudiante_id=estudiante_id,
            materia=materia,
            periodo=periodo,
            nota=nota,
            observaciones=observaciones
        )

        # Crear notificación
        Notificacion.objects.create(
            estudiante_id=estudiante_id,
            tipo='calificacion',
            titulo=f'Nueva calificación en {materia.nombre}',
            mensaje=f'Has recibido una calificación de {nota} en {materia.nombre} - {calificacion.get_periodo_display()}'
        )

        calificacion.notificado = True
        calificacion.save()

        messages.success(request, 'Calificación registrada exitosamente')
        return redirect('teacher_calificaciones_lista')

    # Obtener estudiantes de las materias del docente
    cursos_ids = materias.values_list('curso_id', flat=True).distinct()
    estudiantes = CustomUser.objects.filter(
        role='estudiante',
        is_active=True,
        matriculas__curso_id__in=cursos_ids,
        matriculas__activa=True
    ).distinct()

    context = {
        'materias': materias,
        'estudiantes': estudiantes,
        'periodos': Calificacion.PERIODO_CHOICES,
    }
    return render(request, 'teacher/calificacion_form.html', context)


@login_required
@user_passes_test(is_teacher)
def calificacion_editar(request, calificacion_id):
    """Editar calificación"""
    calificacion = get_object_or_404(Calificacion, id=calificacion_id, materia__docente=request.user)

    if request.method == 'POST':
        calificacion.nota = request.POST.get('nota')
        calificacion.observaciones = request.POST.get('observaciones', '')
        calificacion.save()

        messages.success(request, 'Calificación actualizada exitosamente')
        return redirect('teacher_calificaciones_lista')

    context = {
        'calificacion': calificacion,
        'edit_mode': True,
    }
    return render(request, 'teacher/calificacion_form.html', context)


@login_required
@user_passes_test(is_teacher)
def calificacion_eliminar(request, calificacion_id):
    """Eliminar calificación"""
    calificacion = get_object_or_404(Calificacion, id=calificacion_id, materia__docente=request.user)

    if request.method == 'POST':
        calificacion.delete()
        messages.success(request, 'Calificación eliminada')
        return redirect('teacher_calificaciones_lista')

    return render(request, 'teacher/calificacion_confirmar_eliminar.html', {'calificacion': calificacion})


# ==================== GESTIÓN DE ASISTENCIA ====================

@login_required
@user_passes_test(is_teacher)
def asistencias_lista(request):
    """Lista de asistencias del docente"""
    materias = Materia.objects.filter(docente=request.user, activa=True)
    asistencias = Asistencia.objects.filter(materia__in=materias).select_related(
        'estudiante', 'materia'
    ).order_by('-fecha')

    # Filtros
    materia_id = request.GET.get('materia')
    estado = request.GET.get('estado')
    fecha = request.GET.get('fecha')

    if materia_id:
        asistencias = asistencias.filter(materia_id=materia_id)
    if estado:
        asistencias = asistencias.filter(estado=estado)
    if fecha:
        asistencias = asistencias.filter(fecha=fecha)

    context = {
        'asistencias': asistencias,
        'materias': materias,
        'estados': Asistencia.ESTADO_CHOICES,
        'materia_filter': materia_id,
        'estado_filter': estado,
        'fecha_filter': fecha,
    }
    return render(request, 'teacher/asistencias_lista.html', context)


@login_required
@user_passes_test(is_teacher)
def asistencia_crear(request):
    """Registrar asistencia"""
    materias = Materia.objects.filter(docente=request.user, activa=True)

    if request.method == 'POST':
        estudiante_id = request.POST.get('estudiante')
        materia_id = request.POST.get('materia')
        fecha = request.POST.get('fecha')
        estado = request.POST.get('estado')
        observaciones = request.POST.get('observaciones', '')

        # Verificar que la materia pertenece al docente
        materia = get_object_or_404(Materia, id=materia_id, docente=request.user)

        # Verificar si ya existe un registro
        if Asistencia.objects.filter(estudiante_id=estudiante_id, materia=materia, fecha=fecha).exists():
            messages.error(request, 'Ya existe un registro de asistencia para este estudiante en esta fecha')
            return redirect('teacher_asistencia_crear')

        # Crear asistencia
        Asistencia.objects.create(
            estudiante_id=estudiante_id,
            materia=materia,
            fecha=fecha,
            estado=estado,
            observaciones=observaciones,
            registrado_por=request.user
        )

        messages.success(request, 'Asistencia registrada exitosamente')
        return redirect('teacher_asistencias_lista')

    # Obtener estudiantes de las materias del docente
    cursos_ids = materias.values_list('curso_id', flat=True).distinct()
    estudiantes = CustomUser.objects.filter(
        role='estudiante',
        is_active=True,
        matriculas__curso_id__in=cursos_ids,
        matriculas__activa=True
    ).distinct()

    context = {
        'materias': materias,
        'estudiantes': estudiantes,
        'estados': Asistencia.ESTADO_CHOICES,
        'fecha_hoy': timezone.now().date(),
    }
    return render(request, 'teacher/asistencia_form.html', context)


@login_required
@user_passes_test(is_teacher)
def asistencia_editar(request, asistencia_id):
    """Editar asistencia"""
    asistencia = get_object_or_404(Asistencia, id=asistencia_id, materia__docente=request.user)

    if request.method == 'POST':
        asistencia.estado = request.POST.get('estado')
        asistencia.observaciones = request.POST.get('observaciones', '')
        asistencia.save()

        messages.success(request, 'Asistencia actualizada')
        return redirect('teacher_asistencias_lista')

    context = {
        'asistencia': asistencia,
        'estados': Asistencia.ESTADO_CHOICES,
        'edit_mode': True,
    }
    return render(request, 'teacher/asistencia_form.html', context)


@login_required
@user_passes_test(is_teacher)
def asistencia_eliminar(request, asistencia_id):
    """Eliminar asistencia"""
    asistencia = get_object_or_404(Asistencia, id=asistencia_id, materia__docente=request.user)

    if request.method == 'POST':
        asistencia.delete()
        messages.success(request, 'Asistencia eliminada')
        return redirect('teacher_asistencias_lista')

    return render(request, 'teacher/asistencia_confirmar_eliminar.html', {'asistencia': asistencia})


# ==================== REPORTES Y ESTADÍSTICAS ====================

@login_required
@user_passes_test(is_teacher)
def generar_reporte(request):
    """Generar reporte de calificaciones en Excel"""
    materia_id = request.GET.get('materia')

    if not materia_id:
        messages.error(request, 'Debes seleccionar una materia')
        return redirect('teacher_dashboard')

    materia = get_object_or_404(Materia, id=materia_id, docente=request.user)

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Calificaciones {materia.codigo}"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Encabezados
    headers = ['Estudiante', 'Periodo', 'Nota', 'Estado', 'Observaciones', 'Fecha']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Datos
    calificaciones = Calificacion.objects.filter(materia=materia).select_related('estudiante').order_by('estudiante__username', 'periodo')

    for row, cal in enumerate(calificaciones, 2):
        ws.cell(row=row, column=1, value=cal.estudiante.get_full_name() or cal.estudiante.username)
        ws.cell(row=row, column=2, value=cal.get_periodo_display())
        ws.cell(row=row, column=3, value=float(cal.nota))
        ws.cell(row=row, column=4, value='Aprobado' if cal.aprobado else 'Reprobado')
        ws.cell(row=row, column=5, value=cal.observaciones)
        ws.cell(row=row, column=6, value=cal.fecha_registro.strftime('%d/%m/%Y'))

    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_{materia.codigo}.xlsx'
    wb.save(response)

    return response


@login_required
@user_passes_test(is_teacher)
def estadisticas(request):
    """Ver estadísticas de las materias del docente"""
    materias = Materia.objects.filter(docente=request.user, activa=True)

    stats = []
    for materia in materias:
        calificaciones = Calificacion.objects.filter(materia=materia)
        total_cal = calificaciones.count()
        promedio = calificaciones.aggregate(Avg('nota'))['nota__avg'] or 0
        aprobados = calificaciones.filter(nota__gte=3.0).count()
        reprobados = calificaciones.filter(nota__lt=3.0).count()

        asistencias = Asistencia.objects.filter(materia=materia)
        total_asist = asistencias.count()
        presentes = asistencias.filter(estado='presente').count()
        porcentaje_asist = (presentes / total_asist * 100) if total_asist > 0 else 0

        stats.append({
            'materia': materia,
            'total_calificaciones': total_cal,
            'promedio': round(promedio, 2),
            'aprobados': aprobados,
            'reprobados': reprobados,
            'porcentaje_asistencia': round(porcentaje_asist, 1),
        })

    return render(request, 'teacher/estadisticas.html', {'stats': stats})
