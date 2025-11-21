"""
Vistas personalizadas para el panel de estudiantes.
Contiene vistas de consulta (solo lectura) y acciones específicas (exportar, marcar notificaciones).
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Avg
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
# Importación COMPLETA de modelos para el contexto del estudiante
from .models import Matricula, Calificacion, Asistencia, Notificacion, InscripcionMateria 


def is_student(user):
    """Verifica si el usuario es un estudiante."""
    return user.role == 'estudiante'


@login_required
@user_passes_test(is_student)
def student_dashboard(request):
    """Dashboard principal para estudiantes, mostrando estadísticas clave."""
    
    # Matrículas activas (Curso)
    matriculas = Matricula.objects.filter(
        estudiante=request.user,
        activa=True
    ).select_related('curso')

    # Materias inscritas directamente (InscripcionMateria)
    materias_inscritas = InscripcionMateria.objects.filter(
        estudiante=request.user
    ).select_related('materia', 'materia__curso')

    # Calificaciones
    calificaciones = Calificacion.objects.filter(
        estudiante=request.user
    ).select_related('materia', 'materia__curso').order_by('-fecha_registro')

    # Promedio general
    promedio = calificaciones.aggregate(Avg('nota'))['nota__avg'] or 0

    # Notificaciones no leídas
    notificaciones = Notificacion.objects.filter(
        estudiante=request.user,
        leida=False
    ).order_by('-creada_en')[:5]

    # Asistencias del mes
    mes_actual = timezone.now().month
    asistencias_mes = Asistencia.objects.filter(
        estudiante=request.user,
        fecha__month=mes_actual
    )
    total_asistencias = asistencias_mes.count()
    presentes = asistencias_mes.filter(estado='presente').count()
    porcentaje_asistencia = (presentes / total_asistencias * 100) if total_asistencias > 0 else 0

    context = {
        'matriculas': matriculas,
        'materias_inscritas': materias_inscritas, 
        'calificaciones': calificaciones,
        'promedio': round(promedio, 2),
        'notificaciones': notificaciones,
        'porcentaje_asistencia': round(porcentaje_asistencia, 1),
    }

    return render(request, 'student/dashboard.html', context)


@login_required
@user_passes_test(is_student)
def mis_calificaciones(request):
    """Ver todas las calificaciones agrupadas por materia."""
    calificaciones = Calificacion.objects.filter(
        estudiante=request.user
    ).select_related('materia', 'materia__curso').order_by('materia__curso', 'materia', 'periodo')

    # Agrupar por materia y calcular promedios
    calificaciones_por_materia = {}
    for cal in calificaciones:
        materia_key = cal.materia.id
        if materia_key not in calificaciones_por_materia:
            calificaciones_por_materia[materia_key] = {
                'materia': cal.materia,
                'calificaciones': [],
                'promedio': 0
            }
        calificaciones_por_materia[materia_key]['calificaciones'].append(cal)

    for materia_data in calificaciones_por_materia.values():
        notas = [float(cal.nota) for cal in materia_data['calificaciones'] if cal.nota is not None]
        materia_data['promedio'] = round(sum(notas) / len(notas), 2) if notas else 0

    context = {
        'calificaciones_por_materia': calificaciones_por_materia,
    }
    return render(request, 'student/calificaciones.html', context)


@login_required
@user_passes_test(is_student)
def mis_cursos(request):
    """Ver cursos matriculados y materias inscritas directamente."""
    # Cursos a los que está matriculado
    matriculas = Matricula.objects.filter(
        estudiante=request.user
    ).select_related('curso').order_by('-fecha_matricula')
    
    # Materias a las que está inscrito directamente
    materias_inscritas = InscripcionMateria.objects.filter(
        estudiante=request.user
    ).select_related('materia', 'materia__curso').order_by('materia__curso', 'materia__nombre')

    context = {
        'matriculas': matriculas,
        'materias_inscritas': materias_inscritas,
    }
    return render(request, 'student/cursos.html', context)


@login_required
@user_passes_test(is_student)
def mis_asistencias(request):
    """Ver registro de asistencias y estadísticas."""
    asistencias = Asistencia.objects.filter(
        estudiante=request.user
    ).select_related('materia').order_by('-fecha')

    # Estadísticas
    total = asistencias.count()
    presentes = asistencias.filter(estado='presente').count()
    # Las otras cuentas se mantienen, pero solo presentes afecta el porcentaje
    # ausentes = asistencias.filter(estado='ausente').count()
    # tardanzas = asistencias.filter(estado='tardanza').count()
    # excusados = asistencias.filter(estado='excusado').count()

    porcentaje_asistencia = (presentes / total * 100) if total > 0 else 0

    context = {
        'asistencias': asistencias,
        'total': total,
        'presentes': presentes,
        'ausentes': asistencias.filter(estado='ausente').count(),
        'tardanzas': asistencias.filter(estado='tardanza').count(),
        'excusados': asistencias.filter(estado='excusado').count(),
        'porcentaje_asistencia': round(porcentaje_asistencia, 1),
    }
    return render(request, 'student/asistencias.html', context)


@login_required
@user_passes_test(is_student)
def exportar_calificaciones(request):
    """Exportar calificaciones del estudiante a un archivo Excel."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Mis Calificaciones"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Encabezados
    headers = ['Materia', 'Curso', 'Periodo', 'Nota', 'Observaciones', 'Fecha']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Datos
    calificaciones = Calificacion.objects.filter(
        estudiante=request.user
    ).select_related('materia', 'materia__curso').order_by('materia__curso', 'materia', 'periodo')

    for row, cal in enumerate(calificaciones, 2):
        ws.cell(row=row, column=1, value=cal.materia.nombre)
        ws.cell(row=row, column=2, value=cal.materia.curso.nombre)
        ws.cell(row=row, column=3, value=cal.get_periodo_display())
        ws.cell(row=row, column=4, value=float(cal.nota) if cal.nota is not None else '') 
        ws.cell(row=row, column=5, value=cal.observaciones)
        ws.cell(row=row, column=6, value=cal.fecha_registro.strftime('%d/%m/%Y'))

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=mis_calificaciones_{request.user.username}.xlsx'
    wb.save(response)

    return response


@login_required
@user_passes_test(is_student)
def marcar_notificacion_leida(request, notificacion_id):
    """Marcar notificación específica como leída."""
    notificacion = get_object_or_404(Notificacion, id=notificacion_id, estudiante=request.user)
    notificacion.leida = True
    notificacion.save()
    return redirect('student_dashboard')


@login_required
@user_passes_test(is_student)
def mis_notificaciones(request):
    """Ver todas las notificaciones del estudiante."""
    notificaciones = Notificacion.objects.filter(
        estudiante=request.user
    ).order_by('-creada_en')

    context = {
        'notificaciones': notificaciones,
    }
    return render(request, 'student/notificaciones.html', context)